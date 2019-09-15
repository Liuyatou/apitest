#coding=utf-8

import openpyxl
from Api_excel.common import http_request
from Api_excel.common import contains

class Cases:

    """
    测试用例类 每一个测试用例 实际上就是他的一个实例
    """

    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None
        self.actual=None
        self.result=None
        self.sql=None

class DoExcel:

    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname
        # 打开一个(fileneme)的Excel
        self.workbook=openpyxl.load_workbook(filename)
        # 再定位到sheet
        self.sheet=self.workbook[sheetname]

    def get_cases(self):

        cases=[]   # 列表  存放所以测试用例
        max_row=self.sheet.max_row

        for r in range(2,max_row+1):
            case = Cases() # 实例   每一行对应一个实例
            case.case_id=self.sheet.cell(row=r,column=1).value
            case.title=self.sheet.cell(row=r,column=2).value
            case.url=self.sheet.cell(row=r,column=3).value
            case.data = self.sheet.cell(row=r, column=4).value
            case.method = self.sheet.cell(row=r, column=5).value
            case.expected = self.sheet.cell(row=r, column=6).value
            case.sql = self.sheet.cell(row=r, column=9).value
            cases.append(case)
        return cases # 返回测试列表

    def write_excel(self,row,actual,result):
        sheet=self.workbook[self.sheetname]
        sheet.cell(row,7).value=actual
        sheet.cell(row,8).value=result
        self.workbook.save(self.filename)
        self.workbook.close()

if __name__ == '__main__':

    doexcel=DoExcel(contains.case_file,"login")
    cases=doexcel.get_cases()
    http_request = http_request.HttpRequest_Session()

    for case in cases:
        # print(i.__dict__)
        # print(i.case_id)
        # print(i.title)
        # print(i.url)
        # resp=http_request.request(case.method,case.url,case.data)
        resp=http_request.request(case.method,case.url,case.data)
        actual=resp.text
        print(resp.text)
        resp_dic=resp.json()
        
        print(resp.status_code)
        if case.expected == actual:  # 判断期望结果是否与实际结果一致
            doexcel.write_excel(case.case_id + 1, actual, 'PASS')

        else:
            doexcel.write_excel(case.case_id + 1, actual, 'FAIL')